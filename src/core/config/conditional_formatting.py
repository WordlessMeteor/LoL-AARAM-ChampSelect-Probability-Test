import pandas, openpyxl
from openpyxl.styles import Color, numbers, PatternFill
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule, Rule
from openpyxl.utils import get_column_letter
#声明：每个函数的命名与对应的表头一一对应（Declaration: Each function's naming obeys the one-to-one correspondence with the dataframe header）
def addFormat_LoLHistory_wb(worksheet: openpyxl.workbook.child._WorkbookChild, LoLHistory_df: pandas.DataFrame) -> None:
    #胜负颜色（Win/Lose color）
    col_idx: int = LoLHistory_df.columns.get_loc("result") + 2
    col_letter: str = get_column_letter(col_idx)
    rangeStr: str = "%s3:%s%d" %(col_letter, col_letter, len(LoLHistory_df) + 1)
    win_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "胜利")], stopIfTrue = True, fill = PatternFill(start_color = "63BE7B", end_color = "63BE7B", fill_type = "solid"))
    lose_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "失败")], stopIfTrue = True, fill = PatternFill(start_color = "FF6B6B", end_color = "FF6B6B", fill_type = "solid"))
    terminated_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "被终止")], stopIfTrue = True, fill = PatternFill(start_color = "A6A6A6", end_color = "A6A6A6", fill_type = "solid"))
    worksheet.conditional_formatting.add(rangeStr, win_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, lose_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, terminated_formulaRule_lol)
    #斗魂竞技场队伍排名颜色设置（Arena subteamPlacement color）
    col_idx = LoLHistory_df.columns.get_loc("subteamPlacement") + 2
    col_letter = get_column_letter(col_idx)
    rangeStr = "%s3:%s%d" %(col_letter, col_letter, len(LoLHistory_df) + 1)
    firstPlace_formulaRule_lol: Rule = FormulaRule(formula = ["$%s3=1" %(col_letter)], stopIfTrue = False, fill = PatternFill(start_color = "FFC000", end_color = "FFC000", fill_type = "solid"))
    worksheet.conditional_formatting.add(rangeStr, firstPlace_formulaRule_lol)

def addFormat_LoLGame_info_wb(worksheet: openpyxl.workbook.child._WorkbookChild, LoLGame_info_df: pandas.DataFrame, numColorScale_order: int = 5) -> None:
    #定义条件格式（Define the conditional formats）
    twoDigitPercentage_columns_lol: list[str] = [column for column in LoLGame_info_df.columns if column.endswith("_percent") or column == "GUE"] #百分比（Percentage）
    oneDigitFloat_columns_lol: list[str] = ["KDA"] #一位小数（One-digit float）
    threeDigitFloat_columns_lol: list[str] = ["CSPM", "D/G", "GPM"] #三位小数（Three-digit float）
    colorScale_columns_lol: list[str] = [column for column in LoLGame_info_df.columns if column.endswith("_order")] #条件格式——渐变颜色（Conditional formatting - color scaling）
    dataBar_columns_lol: list[str] = [column for column in LoLGame_info_df.columns if column.endswith("_percent")] #条件格式——数据条（Conditional formatting - data bar）
    order_colorScaleRule_lol: Rule = ColorScaleRule(start_type = "num", start_value = 1, start_color = "63BE7B", mid_type = "percentile", mid_value = 50, mid_color = "FFEB84", end_type = "num", end_value = numColorScale_order, end_color = "FF6B6B") #跳过名次为0的单元格（Skip the order cells whose values are 0）
    percent_dataBarRule_lol: Rule = DataBarRule(start_type = "percentile", start_value = 0, end_type = "percentile", end_value = 100, color = Color("008AEF"), minLength = None, maxLength = None)
    #套用保留两位小数的百分比格式（Two-digit percentage）
    for column in twoDigitPercentage_columns_lol:
        col_idx: int = LoLGame_info_df.columns.get_loc(column) + 2
        for row in range(3, len(LoLGame_info_df) + 2):
            worksheet.cell(row = row, column = col_idx).number_format = numbers.FORMAT_PERCENTAGE_00
    #套用一位小数（One-digit float）
    for column in oneDigitFloat_columns_lol:
        col_idx = LoLGame_info_df.columns.get_loc(column) + 2
        for row in range(3, len(LoLGame_info_df) + 2):
            worksheet.cell(row = row, column = col_idx).number_format = "0.0"
    #套用三位小数（Three-digit float）
    for column in threeDigitFloat_columns_lol:
        col_idx = LoLGame_info_df.columns.get_loc(column) + 2
        for row in range(3, len(LoLGame_info_df) + 2):
            worksheet.cell(row = row, column = col_idx).number_format = "0.000"
    #胜负颜色（Win/Lose color）
    col_idx = LoLGame_info_df.columns.get_loc("win/lose") + 2
    col_letter: str = get_column_letter(col_idx)
    rangeStr: str = "%s3:%s%d" %(col_letter, col_letter, len(LoLGame_info_df) + 1)
    win_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "胜利")], stopIfTrue = True, fill = PatternFill(start_color = "63BE7B", end_color = "63BE7B", fill_type = "solid"))
    lose_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "失败")], stopIfTrue = True, fill = PatternFill(start_color = "FF6B6B", end_color = "FF6B6B", fill_type = "solid"))
    terminated_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "被终止")], stopIfTrue = True, fill = PatternFill(start_color = "A6A6A6", end_color = "A6A6A6", fill_type = "solid"))
    worksheet.conditional_formatting.add(rangeStr, win_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, lose_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, terminated_formulaRule_lol)
    #百分比颜色（Percent color）
    rangeStrs: list[str] = [] #存储尽可能连贯的条件格式区域（Stores continuous conditional formatting areas）
    for i in range(len(dataBar_columns_lol)): #这里需要注意尽量保持条件格式的区域连贯，以免在打开工作簿时条件格式过多导致卡顿（Note that each conditional formatting area should be as large as possible, otherwise the workbook will perform slow when opening it due to too many rules）
        column = dataBar_columns_lol[i]
        if i == 0:
            startCol_idx = endCol_idx = LoLGame_info_df.columns.get_loc(column) + 2
        else:
            col_idx = LoLGame_info_df.columns.get_loc(column) + 2
            if col_idx == endCol_idx + 1: #如果下一个要添加条件格式的列号与上一个要添加条件格式的列号差1，那么这两列是相邻的，即连贯的（If the number of the current column to add conditional format is greater than the number of the predecessive column to add conditional format by 1, then these two columns are continuous）
                endCol_idx = col_idx
            else: #如果两列不相邻，则提取得到上一个连贯的区域（If these two columns aren't continuous, then get the previous continuous area）
                startCol_letter: str = get_column_letter(startCol_idx)
                endCol_letter: str = get_column_letter(endCol_idx)
                rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLGame_info_df) + 1)
                rangeStrs.append(rangeStr)
                startCol_idx = endCol_idx = col_idx #将区域的起始列和终止列设置为当前列（Set the starting and ending columns as the current column）
    else: #执行完成后，把最后一个连贯区域也加上（After the for-loop finishes, add the last continuous area）
        startCol_letter = get_column_letter(startCol_idx)
        endCol_letter = get_column_letter(endCol_idx)
        rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLGame_info_df) + 1)
        rangeStrs.append(rangeStr)
    for rangeStr in rangeStrs:
        worksheet.conditional_formatting.add(rangeStr, percent_dataBarRule_lol)
    #斗魂竞技场队伍排名颜色设置（Arena subteamPlacement color）
    col_idx = LoLGame_info_df.columns.get_loc("subteamPlacement") + 2
    col_letter = get_column_letter(col_idx)
    rangeStr = "%s3:%s%d" %(col_letter, col_letter, len(LoLGame_info_df) + 1)
    firstPlace_formulaRule_lol: Rule = FormulaRule(formula = ["$%s3=1" %(col_letter)], stopIfTrue = False, fill = PatternFill(start_color = "FFC000", end_color = "FFC000", fill_type = "solid"))
    worksheet.conditional_formatting.add(rangeStr, firstPlace_formulaRule_lol)
    #位次颜色（Order color）
    rangeStrs = [] #存储尽可能连贯的条件格式区域（Stores continuous conditional formatting areas）
    rangeTuples: list[tuple[str, str]] = []
    for i in range(len(colorScale_columns_lol)): #这里需要注意尽量保持条件格式的区域连贯，以免在打开工作簿时条件格式过多导致卡顿（Note that each conditional formatting area should be as large as possible, otherwise the workbook will perform slow when opening it due to too many rules）
        column = colorScale_columns_lol[i]
        if i == 0:
            startCol_idx = endCol_idx = LoLGame_info_df.columns.get_loc(column) + 2
        else:
            col_idx = LoLGame_info_df.columns.get_loc(column) + 2
            if col_idx == endCol_idx + 1: #如果下一个要添加条件格式的列号与上一个要添加条件格式的列号差1，那么这两列是相邻的，即连贯的（If the number of the current column to add conditional format is greater than the number of the predecessive column to add conditional format by 1, then these two columns are continuous）
                endCol_idx = col_idx
            else: #如果两列不相邻，则提取得到上一个连贯的区域（If these two columns aren't continuous, then get the previous continuous area）
                startCol_letter = get_column_letter(startCol_idx)
                endCol_letter = get_column_letter(endCol_idx)
                rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLGame_info_df) + 1)
                rangeStrs.append(rangeStr)
                rangeTuples.append((startCol_letter, endCol_letter))
                startCol_idx = endCol_idx = col_idx #将区域的起始列和终止列设置为当前列（Set the starting and ending columns as the current column）
    else: #执行完成后，把最后一个连贯区域也加上（After the for-loop finishes, add the last continuous area）
        startCol_letter = get_column_letter(startCol_idx)
        endCol_letter = get_column_letter(endCol_idx)
        rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLGame_info_df) + 1)
        rangeStrs.append(rangeStr)
        rangeTuples.append((startCol_letter, endCol_letter))
    for i in range(len(rangeStrs)):
        rangeStr = rangeStrs[i]
        rangeTuple = rangeTuples[i]
        order_noFillRule: Rule = FormulaRule(formula = ["%s3=0" %(rangeTuple[0])], stopIfTrue = True, fill = PatternFill(fill_type = None))
        worksheet.conditional_formatting.add(rangeStr, order_noFillRule)
        worksheet.conditional_formatting.add(rangeStr, order_colorScaleRule_lol)

def addFormat_LoLGame_info_wb_transpose(worksheet: openpyxl.workbook.child._WorkbookChild, LoLGame_info_df: pandas.DataFrame, numColorScale_order: int = 5) -> None:
    #定义条件格式（Define the conditional formats）
    twoDigitPercentage_rows_lol: list[str] = [row for row in LoLGame_info_df.index if row.endswith("_percent") or row == "GUE"] #百分比（Percentage）
    oneDigitFloat_rows_lol: list[str] = ["KDA"] #一位小数（One-digit float）
    threeDigitFloat_rows_lol: list[str] = ["CSPM", "D/G", "GPM"] #三位小数（Three-digit float）
    colorScale_rows_lol: list[str] = [row for row in LoLGame_info_df.index if row.endswith("_order")] #条件格式——渐变颜色（Conditional formatting - color scaling）
    dataBar_rows_lol: list[str] = [row for row in LoLGame_info_df.index if row.endswith("_percent")] #条件格式——数据条（Conditional formatting - data bar）
    order_colorScaleRule_lol: Rule = ColorScaleRule(start_type = "num", start_value = 1, start_color = "63BE7B", mid_type = "percentile", mid_value = 50, mid_color = "FFEB84", end_type = "num", end_value = numColorScale_order, end_color = "FF6B6B") #跳过名次为0的单元格（Skip the order cells whose values are 0）
    percent_dataBarRule_lol: Rule = DataBarRule(start_type = "percentile", start_value = 0, end_type = "percentile", end_value = 100, color = Color("008AEF"), minLength = None, maxLength = None)
    #套用保留两位小数的百分比格式（Two-digit percentage）
    for row in twoDigitPercentage_rows_lol:
        row_idx: int = LoLGame_info_df.index.get_loc(row) + 2
        for column in range(3, len(LoLGame_info_df) + 2):
            worksheet.cell(column = column, row = row_idx).number_format = numbers.FORMAT_PERCENTAGE_00
    #套用一位小数（One-digit float）
    for row in oneDigitFloat_rows_lol:
        row_idx = LoLGame_info_df.index.get_loc(row) + 2
        for column in range(3, len(LoLGame_info_df) + 2):
            worksheet.cell(column = column, row = row_idx).number_format = "0.0"
    #套用三位小数（Three-digit float）
    for row in threeDigitFloat_rows_lol:
        row_idx = LoLGame_info_df.index.get_loc(row) + 2
        for column in range(3, len(LoLGame_info_df) + 2):
            worksheet.cell(column = column, row = row_idx).number_format = "0.000"
    #胜负颜色（Win/Lose color）
    row_idx = LoLGame_info_df.index.get_loc("win/lose") + 2
    col_letter: str = get_column_letter(len(LoLGame_info_df) + 1)
    rangeStr: str = "C%d:%s%d" %(row_idx, col_letter, row_idx)
    win_formulaRule_lol: Rule = FormulaRule(formula = ['C$%d="%s"' %(row_idx, "胜利")], stopIfTrue = True, fill = PatternFill(start_color = "63BE7B", end_color = "63BE7B", fill_type = "solid"))
    lose_formulaRule_lol: Rule = FormulaRule(formula = ['C$%d="%s"' %(row_idx, "失败")], stopIfTrue = True, fill = PatternFill(start_color = "FF6B6B", end_color = "FF6B6B", fill_type = "solid"))
    terminated_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "被终止")], stopIfTrue = True, fill = PatternFill(start_color = "A6A6A6", end_color = "A6A6A6", fill_type = "solid"))
    worksheet.conditional_formatting.add(rangeStr, win_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, lose_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, terminated_formulaRule_lol)
    #百分比颜色（Percent color）
    rangeStrs: list[str] = [] #存储尽可能连贯的条件格式区域（Stores continuous conditional formatting areas）
    for i in range(len(dataBar_rows_lol)): #这里需要注意尽量保持条件格式的区域连贯，以免在打开工作簿时条件格式过多导致卡顿（Note that each conditional formatting area should be as large as possible, otherwise the workbook will perform slow when opening it due to too many rules）
        row = dataBar_rows_lol[i]
        if i == 0:
            startRow_idx = endRow_idx = LoLGame_info_df.index.get_loc(row) + 2
        else:
            row_idx = LoLGame_info_df.index.get_loc(row) + 2
            if row_idx == endRow_idx + 1: #如果下一个要添加条件格式的行号与上一个要添加条件格式的行号差1，那么这两行是相邻的，即连贯的（If the number of the current row to add conditional format is greater than the number of the predecessive row to add conditional format by 1, then these two rows are continuous）
                endRow_idx = row_idx
            else: #如果两行不相邻，则提取得到上一个连贯的区域（If these two rows aren't continuous, then get the previous continuous area）
                endCol_letter: str = get_column_letter(len(LoLGame_info_df) + 1)
                rangeStr = "C%d:%s%d" %(startRow_idx, endCol_letter, endRow_idx)
                rangeStrs.append(rangeStr)
                startRow_idx = endRow_idx = row_idx #将区域的起始行和终止行设置为当前行（Set the starting and ending rows as the current row）
    else: #执行完成后，把最后一个连贯区域也加上（After the for-loop finishes, add the last continuous area）
        endCol_letter = get_column_letter(len(LoLGame_info_df) + 1)
        rangeStr = "C%d:%s%d" %(startRow_idx, endCol_letter, endRow_idx)
        rangeStrs.append(rangeStr)
    for rangeStr in rangeStrs:
        worksheet.conditional_formatting.add(rangeStr, percent_dataBarRule_lol)
    #斗魂竞技场队伍排名颜色设置（Arena subteamPlacement color）
    row_idx = LoLGame_info_df.index.get_loc("subteamPlacement") + 2
    col_letter = get_column_letter(len(LoLGame_info_df) + 1)
    rangeStr = "C%d:%s%d" %(row_idx, col_letter, row_idx)
    firstPlace_formulaRule_lol: Rule = FormulaRule(formula = ["C$%d=1" %(row_idx)], stopIfTrue = False, fill = PatternFill(start_color = "FFC000", end_color = "FFC000", fill_type = "solid"))
    worksheet.conditional_formatting.add(rangeStr, firstPlace_formulaRule_lol)
    #位次颜色（Order color）
    rangeStrs = [] #存储尽可能连贯的条件格式区域（Stores continuous conditional formatting areas）
    rangeTuples: list[tuple[int, int]] = []
    for i in range(len(colorScale_rows_lol)): #这里需要注意尽量保持条件格式的区域连贯，以免在打开工作簿时条件格式过多导致卡顿（Note that each conditional formatting area should be as large as possible, otherwise the workbook will perform slow when opening it due to too many rules）
        row = colorScale_rows_lol[i]
        if i == 0:
            startRow_idx = endRow_idx = LoLGame_info_df.index.get_loc(row) + 2
        else:
            row_idx = LoLGame_info_df.index.get_loc(row) + 2
            if row_idx == endRow_idx + 1: #如果下一个要添加条件格式的行号与上一个要添加条件格式的行号差1，那么这两行是相邻的，即连贯的（If the number of the current row to add conditional format is greater than the number of the predecessive row to add conditional format by 1, then these two rows are continuous）
                endRow_idx = row_idx
            else: #如果两列不相邻，则提取得到上一个连贯的区域（If these two columns aren't continuous, then get the previous continuous area）
                endCol_letter = get_column_letter(len(LoLGame_info_df) + 1)
                rangeStr = "C%d:%s%d" %(startRow_idx, endCol_letter, endRow_idx)
                rangeStrs.append(rangeStr)
                rangeTuples.append((startRow_idx, endRow_idx))
                startRow_idx = endRow_idx = row_idx #将区域的起始行和终止行设置为当前行（Set the starting and ending rows as the current row）
    else: #执行完成后，把最后一个连贯区域也加上（After the for-loop finishes, add the last continuous area）
        endCol_letter = get_column_letter(len(LoLGame_info_df) + 1)
        rangeStr = "C%d:%s%d" %(startRow_idx, endCol_letter, endRow_idx)
        rangeStrs.append(rangeStr)
        rangeTuples.append((startRow_idx, endRow_idx))
    for i in range(len(rangeStrs)):
        rangeStr = rangeStrs[i]
        rangeTuple = rangeTuples[i]
        order_noFillRule: Rule = FormulaRule(formula = ["C%d=0" %(rangeTuple[0])], stopIfTrue = True, fill = PatternFill(fill_type = None))
        worksheet.conditional_formatting.add(rangeStr, order_noFillRule)
        worksheet.conditional_formatting.add(rangeStr, order_colorScaleRule_lol)

def addFormat_LoLPlayer_summary_wb(worksheet: openpyxl.workbook.child._WorkbookChild, LoLPlayer_summary_df: pandas.DataFrame, numColorScale_order: int = 5) -> None:
    #定义条件格式（Define the conditional formats）
    twoDigitPercentage_columns_lol_summary: list[str] = ["KP_percent"] #百分比（Percentage）
    oneDigitFloat_columns_lol_summary: list[str] = ["KDA"] #一位小数（One-digit float）
    colorScale_columns_lol_summary: list[str] = [column for column in LoLPlayer_summary_df.columns if column.endswith("_order")] #条件格式——渐变颜色（Conditional formatting - color scaling）
    dataBar_columns_lol_summary: list[str] = [column for column in LoLPlayer_summary_df.columns if column.endswith("_percent")] #条件格式——数据条（Conditional formatting - data bar）
    order_colorScaleRule_lol: Rule = ColorScaleRule(start_type = "num", start_value = 1, start_color = "63BE7B", mid_type = "percentile", mid_value = 50, mid_color = "FFEB84", end_type = "num", end_value = numColorScale_order, end_color = "FF6B6B") #跳过名次为0的单元格。这里`end_value`的选取可以讨论一下，可以选取所有对局的队列信息中记录的队伍规模的最大值（Skip the order cells whose values are 0. Here the value of `end_value` is worth discussion: it may take the maximum of `numPlayersPerTeam` recorded in the queue data of the corresponding queueIds）
    percent_dataBarRule_lol: Rule = DataBarRule(start_type = "percentile", start_value = 0, end_type = "percentile", end_value = 100, color = Color("008AEF"), minLength = None, maxLength = None)
    #套用保留两位小数的百分比格式（Two-digit percentage）
    for column in twoDigitPercentage_columns_lol_summary:
        col_idx: int = LoLPlayer_summary_df.columns.get_loc(column) + 2 #Excel中的第一列（A列）的索引是1，且又是数据框的索引列【The index of the first column (Column A) in Excel is 1, and this column is the index of column of the dataframe）
        for row in range(3, len(LoLPlayer_summary_df) + 2):
            worksheet.cell(row = row, column = col_idx).number_format = numbers.FORMAT_PERCENTAGE_00
    #套用一位小数（One-digit float）
    for column in oneDigitFloat_columns_lol_summary:
        col_idx = LoLPlayer_summary_df.columns.get_loc(column) + 2
        for row in range(3, len(LoLPlayer_summary_df) + 2):
            worksheet.cell(row = row, column = col_idx).number_format = "0.0"
    #胜负颜色（Win/Lose color）
    col_idx = LoLPlayer_summary_df.columns.get_loc("win/lose") + 2
    col_letter: str = get_column_letter(col_idx)
    rangeStr: str = "%s3:%s%d" %(col_letter, col_letter, len(LoLPlayer_summary_df) + 1)
    win_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "胜利")], stopIfTrue = True, fill = PatternFill(start_color = "63BE7B", end_color = "63BE7B", fill_type = "solid"))
    lose_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "失败")], stopIfTrue = True, fill = PatternFill(start_color = "FF6B6B", end_color = "FF6B6B", fill_type = "solid"))
    terminated_formulaRule_lol: Rule = FormulaRule(formula = ['$%s3="%s"' %(col_letter, "被终止")], stopIfTrue = True, fill = PatternFill(start_color = "A6A6A6", end_color = "A6A6A6", fill_type = "solid"))
    worksheet.conditional_formatting.add(rangeStr, win_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, lose_formulaRule_lol)
    worksheet.conditional_formatting.add(rangeStr, terminated_formulaRule_lol)
    #百分比颜色（Percent color）
    rangeStrs: list[str] = [] #存储尽可能连贯的条件格式区域（Stores continuous conditional formatting areas）
    for i in range(len(dataBar_columns_lol_summary)): #这里需要注意尽量保持条件格式的区域连贯，以免在打开工作簿时条件格式过多导致卡顿（Note that each conditional formatting area should be as large as possible, otherwise the workbook will perform slow when opening it due to too many rules）
        column: list[str] = dataBar_columns_lol_summary[i]
        if i == 0:
            startCol_idx = endCol_idx = LoLPlayer_summary_df.columns.get_loc(column) + 2
        else:
            col_idx = LoLPlayer_summary_df.columns.get_loc(column) + 2
            if col_idx == endCol_idx + 1: #如果下一个要添加条件格式的列号与上一个要添加条件格式的列号差1，那么这两列是相邻的，即连贯的（If the number of the current column to add conditional format is greater than the number of the predecessive column to add conditional format by 1, then these two columns are continuous）
                endCol_idx = col_idx
            else: #如果两列不相邻，则提取得到上一个连贯的区域（If these two columns aren't continuous, then get the previous continuous area）
                startCol_letter: str = get_column_letter(startCol_idx)
                endCol_letter: str = get_column_letter(endCol_idx)
                rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLPlayer_summary_df) + 1)
                rangeStrs.append(rangeStr)
                startCol_idx = endCol_idx = col_idx #将区域的起始列和终止列设置为当前列（Set the starting and ending columns as the current column）
    else: #执行完成后，把最后一个连贯区域也加上（After the for-loop finishes, add the last continuous area）
        startCol_letter = get_column_letter(startCol_idx)
        endCol_letter = get_column_letter(endCol_idx)
        rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLPlayer_summary_df) + 1)
        rangeStrs.append(rangeStr)
    for rangeStr in rangeStrs:
        worksheet.conditional_formatting.add(rangeStr, percent_dataBarRule_lol)
    #位次颜色（Order color）
    rangeStrs = []
    rangeTuples: list[tuple[int, int]] = []
    for i in range(len(colorScale_columns_lol_summary)):
        column = colorScale_columns_lol_summary[i]
        if i == 0:
            startCol_idx = endCol_idx = LoLPlayer_summary_df.columns.get_loc(column) + 2
        else:
            col_idx = LoLPlayer_summary_df.columns.get_loc(column) + 2
            if col_idx == endCol_idx + 1:
                endCol_idx = col_idx
            else:
                startCol_letter = get_column_letter(startCol_idx)
                endCol_letter = get_column_letter(endCol_idx)
                rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLPlayer_summary_df) + 1)
                rangeStrs.append(rangeStr)
                rangeTuples.append((startCol_letter, endCol_letter))
                startCol_idx = endCol_idx = col_idx
    else:
        startCol_letter = get_column_letter(startCol_idx)
        endCol_letter = get_column_letter(endCol_idx)
        rangeStr = "%s3:%s%d" %(startCol_letter, endCol_letter, len(LoLPlayer_summary_df) + 1)
        rangeStrs.append(rangeStr)
        rangeTuples.append((startCol_letter, endCol_letter))
    for i in range(len(rangeStrs)):
        rangeStr = rangeStrs[i]
        rangeTuple = rangeTuples[i]
        order_noFillRule: Rule = FormulaRule(formula = ["%s3=0" %(rangeTuple[0])], stopIfTrue = True, fill = PatternFill(fill_type = None))
        worksheet.conditional_formatting.add(rangeStr, order_noFillRule)
        worksheet.conditional_formatting.add(rangeStr, order_colorScaleRule_lol)

def addFormat_inGame_allPlayer_wb(worksheet: openpyxl.workbook.child._WorkbookChild, inGame_allPlayer_df: pandas.DataFrame) -> None:
    #定义条件格式（Define the conditional formats）
    oneDigitFloat_columns_lol: list[str] = ["KDA"] #一位小数（One-digit float）
    threeDigitFloat_columns_lol: list[str] = ["CSPM"] #三位小数（Three-digit float）
    #套用一位小数（One-digit float）
    for column in oneDigitFloat_columns_lol:
        col_idx: int = inGame_allPlayer_df.columns.get_loc(column) + 2
        for row in range(3, len(inGame_allPlayer_df) + 2):
            worksheet.cell(row = row, column = col_idx).number_format = "0.0"
    #套用三位小数（Three-digit float）
    for column in threeDigitFloat_columns_lol:
        col_idx = inGame_allPlayer_df.columns.get_loc(column) + 2
        for row in range(3, len(inGame_allPlayer_df) + 2):
            worksheet.cell(row = row, column = col_idx).number_format = "0.000"
